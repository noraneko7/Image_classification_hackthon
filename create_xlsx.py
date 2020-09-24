Excelファイル生成
import openpyxl
import os
import shutil
import glob
import re
import pandas as pd
import json
#読み込みたいcsv
csv = '/Teddy Bear_end.csv'
#フォルダー番号（企業）
num = '0'
#フォルダー名（ジャンル）
genre = 'Teddy Bear'
#正解クラス
true_class = genre
#csvファイルをインポート
df = pd.read_csv(csv)
class_list = []
class_name = []
tmp = []
count = 0
for index, row in df.iterrows():
  count += 1
  point = row['index'].find('/00')+1
  if row['class'] != true_class:
    tmp.append(row['index'][point:])
    class_list.append(row['index'][point:])
tmp.sort()
class_list.sort()
# ワークブックを新規作成する
book = openpyxl.Workbook()
# シートを取得し名前を変更する
sheet = book.active
sheet.title = 'VGG16分類結果.'+num+'.'+genre
list_a = ['企業番号', 'クラス名', '', '正答率', '間違い数', '写真数', '間違い写真番号']
i = 1
for item in list_a:
  sheet.cell(row=i, column=1).value = item
  i += 1
list_b = ['人力判定', 'VGG16',  '画像の種類', '分類されたクラス']
i = 2
for item in list_b:
  sheet.cell(row=3, column=i).value = item
  i += 1
sheet.cell(row=1, column=2).value = 'フォルダー'+num
sheet.cell(row=2, column=2).value = genre
sheet.cell(row=2, column=3).value = '正解クラス→'+true_class
sheet.cell(row=6, column=2).value = count
sheet.cell(row=6, column=3).value = count
#↓データセット↓
#全ての不正解データ
list_1 = tmp
#自力の不正解データ
json_open = open('/content/drive/My Drive/Colab Notebooks/data_list_v3.json', 'r')
json_load = json.load(json_open)
for list_num in list(json_load.keys()):
  if list_num == num:
    all_obj = json_load[list_num]
for obj_name in all_obj.keys():
  if obj_name == genre:
     list_2 = all_obj[obj_name]
#VGG16による不正解判定のデータ
list_3 = class_list
i=0
j=0
x=0
sell = 7
sheet.cell(row=5, column=2).value = len(list_2)
sheet.cell(row=5, column=3).value = len(list_3)
sheet.cell(row=4, column=2).value = '=ROUND(100 -(B5/B6 * 100),2)'
sheet.cell(row=4, column=3).value = '=ROUND(100 -(C5/C6 * 100),2)'
#結合
list_1.extend(list_2)
#ソート
list_1.sort()
#重複削除
list_1 = dict.fromkeys(list_1)
for img in list(list_1):
  for index, row in df.iterrows():
    point = row['index'].find('/00')+1
    if row['index'][point:] == img:
      class_name.append(row['class'])
for item in list(list_1):
    if item == list_2[i]:
        if item == list_3[j]:
            sheet.cell(row=sell, column=2).value = item
            sheet.cell(row=sell, column=3).value = item
            sheet.cell(row=sell, column=5).value = class_name[x]
            sell += 1
            x += 1
            if i == len(list_2)-1:
              if j == len(list_3)-1:
                j += 1
            elif j == len(list_3)-1:
                i += 1
            else:
                i += 1
                j += 1
        else:
            sheet.cell(row=sell, column=2).value = item
            sheet.cell(row=sell, column=5).value = class_name[x]
            sell += 1
            x += 1
            if i != len(list_2)-1:
                i += 1
    elif item == list_3[j]:
        sheet.cell(row=sell, column=3).value = item
        sheet.cell(row=sell, column=5).value = class_name[x]
        sell += 1
        x += 1
        if j != len(list_3)-1:
            j += 1
sheet.column_dimensions["A"].width = 18
sheet.column_dimensions["B"].width = 12
sheet.column_dimensions["C"].width = 12
sheet.column_dimensions["D"].width = 20
sheet.column_dimensions["E"].width = 28
book.save('vgg16_'+num+'_'+true_class+'.xlsx')